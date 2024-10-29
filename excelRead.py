from ttkbootstrap import Style, Button, OptionMenu
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter.scrolledtext import ScrolledText
import random
from PIL import Image, ImageTk
import openpyxl
import pandas as pd

root = tk.Tk()
style = Style(theme='lumen')

root.title("Excel 自动随机选号助手")
root.geometry("300x400")

# 使用 PIL 库加载图像并转换为 Tkinter 可以使用的格式
icon_path = "./img/2.png"
root.iconphoto(False, ImageTk.PhotoImage(Image.open(icon_path)))

# 设置字体
# font_family = "STXingkai"
# font_size = 10

# 定义按钮样式
button_style = 'primary.Outline.TButton'
style.configure(button_style, font=('Helvetica', 10), borderwidth=2, relief='raised', padding=(10, 5), radius=8)

# 使用 PIL 库加载图像并转换为 Tkinter 可以使用的格式
image_path = "./img/1.png"  # 替换为你的图像文件路径
img = Image.open(image_path)
bg_image = ImageTk.PhotoImage(img)

bg_label = tk.Label(root, image=bg_image)
bg_label.place(x=0, y=0, relwidth=1, relheight=1)

# 使用 ttkbootstrap 的 Button
file_button = Button(root, text="选择 Excel 文件", style=button_style)
file_button.pack(pady=10)

sheet_var = tk.StringVar(root)
# 使用 ttkbootstrap 的 OptionMenu
sheet_dropdown = OptionMenu(root, sheet_var, "", style='TMenubutton')
sheet_dropdown.pack(pady=5)

column_var = tk.StringVar(root)
column_dropdown = OptionMenu(root, column_var, "", style='TMenubutton')
column_dropdown.pack(pady=5)

unique_data_var = tk.StringVar(root)
unique_data_dropdown = OptionMenu(root, unique_data_var, "", style='TMenubutton')
unique_data_dropdown.pack(pady=5)

file_path = None
selected_sheet_name = None

# 创建一个标签用于显示提示信息，靠底部显示，有圆角和浅蓝色填充
info_label = tk.Label(root, text="欢迎使用 Excel 自动随机选号工具,请选择文件", bd=0, relief='flat')
info_label.pack(side=tk.BOTTOM, pady=10)
info_label.config(bg='#ADD8E6', borderwidth=2, relief='ridge')


def load_excel_data():
    global file_path, selected_sheet_name
    # 更新提示信息
    info_label.config(text="选择文件加载中，请等待...")
    file_path = filedialog.askopenfilename(filetypes=[("Excel 文件", "*.xlsx")])
    if file_path:
        # 使用 openpyxl 的 read_only 模式读取 Excel 文件
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = list(wb.sheetnames)
        sheet_var.set("选择 sheet")
        sheet_dropdown['menu'].delete(0, 'end')
        for name in sheet_names:
            def on_sheet_selection(sheet_name_inner=name):
                sheet_var.set(sheet_name_inner)
                # 更新提示信息
                info_label.config(text="Sheet 加载中，请等待")
                on_sheet_select(sheet_name_inner)

            sheet_dropdown['menu'].add_command(label=name, command=on_sheet_selection)
        selected_sheet_name = None
        # 更新提示信息
        info_label.config(text="选择文件操作成功，请选择 Sheet")
    else:
        # 如果未选择文件，也更新提示信息
        info_label.config(text="未选择文件")


def select_file():
    load_excel_data()


def on_sheet_select(sheet_name):
    global file_path, selected_sheet_name, column_names
    # 更新提示信息
    info_label.config(text="Sheet 加载中，请等待")
    if file_path:
        selected_sheet_name = sheet_name
        # 使用 openpyxl 的 read_only 模式读取 Excel 文件
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[sheet_name]
        column_names = [cell.value for cell in ws[1]]
        column_var.set("选择列名")
        column_dropdown['menu'].delete(0, 'end')
        for column in column_names:
            # 更新提示信息
            info_label.config(text="Sheet 加载中，请等待")

            def on_column_selection(column_inner=column):
                column_var.set(column_inner)
                on_column_select(column_inner)

            column_dropdown['menu'].add_command(label=column, command=on_column_selection)
        # 更新提示信息
        info_label.config(text="选择 Sheet操作成功,请选择列")


def on_column_select(selected_column):
    global file_path, selected_sheet_name, column_names
    # 更新提示信息
    info_label.config(text="数据加载中，请等待")
    if file_path:
        # 使用 openpyxl 的 read_only 模式读取 Excel 文件
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[selected_sheet_name]
        unique_data = set()
        for row in ws.iter_rows(values_only=True):
            if row:
                value = row[column_names.index(selected_column)]
                if value is not None:
                    unique_data.add(value)
        unique_data_var.set("选择目标数据")
        # 先清空菜单
        unique_data_dropdown['menu'].delete(0, 'end')
        # 创建搜索框和关联的变量
        search_var = tk.StringVar()
        search_entry = tk.Entry(root, textvariable=search_var)
        search_entry.pack(pady=5)

        def filter_unique_data(*args):
            search_text = search_var.get().lower()
            filtered_data = [data for data in unique_data if search_text in str(data).lower()]
            unique_data_dropdown['menu'].delete(0, 'end')
            for data in filtered_data:
                def on_unique_data_selection(data_inner=data):
                    unique_data_var.set(data_inner)
                    show_rows_with_selected_data(selected_column, data_inner)

                unique_data_dropdown['menu'].add_command(label=data, command=on_unique_data_selection)

        search_var.trace_add("write", filter_unique_data)
        for data in unique_data:
            def on_unique_data_selection(data_inner=data):
                unique_data_var.set(data_inner)
                show_rows_with_selected_data(selected_column, data_inner)

            unique_data_dropdown['menu'].add_command(label=data, command=on_unique_data_selection)
        # 更新提示信息
        info_label.config(text="选择列操作成功，请选择目标数据")


def show_rows_with_selected_data(selected_column, selected_data):
    global file_path, selected_sheet_name, column_names
    if file_path:
        # 使用 openpyxl 的 read_only 模式读取 Excel 文件
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[selected_sheet_name]
        selected_rows = []
        for row in ws.iter_rows(values_only=True):
            if row[column_names.index(selected_column)] == selected_data:
                selected_rows.append(row)
        new_window = tk.Toplevel(root)
        new_window.title(f"包含 {selected_data} 的行")

        # 主框架用于左右布局
        main_frame = tk.Frame(new_window)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 左侧显示选中数据的行
        left_frame = tk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        text = ScrolledText(left_frame, height=10, width=50)
        text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        text.insert(tk.END, '\n'.join([str(row) for row in selected_rows]))

        # 右侧框架用于显示行数、数字按钮、确认按钮、抽取按钮和导出按钮及结果显示框
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=5, pady=5)

        # 显示有效数据量
        row_count_label = tk.Label(right_frame, text=f"有效数据量：{len(selected_rows)}", font=("STXingkai", 25, "bold"))
        row_count_label.pack(pady=5)

        # 数字按钮框架
        number_frame = tk.Frame(right_frame)
        number_frame.pack(pady=5)
        number_buttons = []
        for i in range(10):
            button = Button(number_frame, text=str(i), width=5)
            style.configure('TButton', font=('Helvetica', 10))
            button.grid(row=i // 3, column=i % 3, padx=5, pady=5)
            number_buttons.append(button)

        selected_numbers = []

        def select_number(number):
            selected_numbers.append(number)
            update_display()

        for button in number_buttons:
            button['command'] = lambda n=int(button['text']): select_number(n)

        display_label = tk.Label(right_frame, text='')
        display_label.pack(pady=5)

        def update_display():
            display_label.config(text=''.join(map(str, selected_numbers)))

        global entry_var

        # 确认按钮在数字按钮的第 4 行第 3 列位置
        def confirm_selection():
            if selected_numbers:
                num_to_extract = int(''.join(map(str, selected_numbers)))
                entry_var.set(num_to_extract)
                selected_numbers.clear()
                update_display()

        entry_var = tk.StringVar()

        confirm_button = Button(number_frame, text="确认", style=button_style, command=confirm_selection)
        confirm_button.grid(row=3, column=2, padx=5, pady=5)

        # 中间显示框
        middle_label = tk.Label(right_frame, textvariable=entry_var, font=("黑体", 20, "bold"))
        middle_label.pack(pady=5)

        # 抽取和导出按钮在一行
        button_frame = tk.Frame(right_frame)
        button_frame.pack(pady=5)
        extract_button = Button(button_frame, text="抽取", style=button_style)
        extract_button.pack(side=tk.LEFT, padx=5)

        def perform_extraction():
            try:
                num_to_extract = int(entry_var.get())
                if num_to_extract <= len(selected_rows):
                    extracted_rows = random.sample(selected_rows, num_to_extract)
                    # 创建一个新的 DataFrame 来存储抽取的行
                    extracted_df = pd.DataFrame(extracted_rows)
                    # 动态调整显示抽取结果的文本框
                    result_text.config(state='normal')
                    result_text.delete(1.0, tk.END)
                    result_text.insert(tk.END, extracted_df.to_string())
                    result_text.config(state='disabled')
                    result_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
                else:
                    messagebox.showerror("错误", "抽取数量不能大于行数。")
            except ValueError:
                messagebox.showerror("错误", "请输入有效的整数。")

        extract_button['command'] = perform_extraction

        export_button = Button(button_frame, text="导出", style=button_style)
        export_button.pack(side=tk.LEFT, padx=5)

        def export_extracted_data():
            try:
                num_to_extract = int(entry_var.get())
                if num_to_extract <= len(selected_rows):
                    extracted_rows = random.sample(selected_rows, num_to_extract)
                    extracted_df = pd.DataFrame(extracted_rows)
                    export_file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                                    filetypes=[("Excel 文件", "*.xlsx")])
                    if export_file_path:
                        extracted_df.to_excel(export_file_path, index=False)
                        messagebox.showinfo("成功", "数据已成功导出。")
                else:
                    messagebox.showerror("错误", "抽取数量不能大于行数。")
            except ValueError:
                messagebox.showerror("错误", "请输入有效的整数。")

        export_button['command'] = export_extracted_data

        # 显示抽取结果的文本框
        result_text = ScrolledText(right_frame, height=10, width=50)
        result_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # 分割线框架
        splitter = tk.Frame(main_frame, width=5, cursor="sb_h_double_arrow")
        splitter.pack(side=tk.LEFT, fill=tk.Y)

        drag_start_width = None

        def start_drag(event):
            nonlocal drag_start_width
            drag_start_width = left_frame.winfo_width()

        def do_drag(event):
            nonlocal drag_start_width
            if drag_start_width is not None:
                delta = event.x
                new_width = drag_start_width + delta
                if 50 <= new_width <= main_frame.winfo_width() - 50:
                    left_frame.config(width=new_width)
                    right_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=5, pady=5, after=left_frame)

        splitter.bind("<ButtonPress-1>", start_drag)
        splitter.bind("<B1-Motion>", do_drag)

        # 分界线
        line_frame = tk.Frame(main_frame, height=main_frame.winfo_height(), width=2)
        line_frame.place(in_=splitter, relx=1, rely=0, anchor='nw')
        line_frame.config(bg='black')


file_button['command'] = select_file

root.mainloop()